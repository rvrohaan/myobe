"""Send the MyOBE 'thank you + institutional plans' email to registered users.

Reads recipient emails from an Excel sheet and sends each person an INDIVIDUAL
email (never BCC) through Resend, with the institutional-plan document attached.

Config (same env vars the web app uses):
    RESEND_API_KEY   your Resend API key (required)
    SMTP_FROM        sender, e.g. "MyOBE <help@myobe.in>" (default provided)
Both are read from the environment or a local .env file.

Usage:
    # 1. Always dry-run first: parses the sheet, prints who WOULD be emailed,
    #    sends nothing.
    python send_bulk_email.py --excel users.xlsx --dry-run

    # 2. Real send once the recipient list looks right:
    python send_bulk_email.py --excel users.xlsx

    # Useful options:
    #   --attachment path.docx   file to attach (default: MyOBE_Institutional_Plans.docx)
    #   --column "Email"         only read emails from this named/indexed column
    #   --sheet "Sheet1"         which worksheet (default: active)
    #   --delay 0.6              seconds between sends (Resend rate limit)
    #   --limit 50               stop after N recipients (handy for a test batch)
    #   --log sent_log.csv       progress log; re-runs SKIP already-sent addresses
    #   --test you@example.com   ignore the sheet, send one email to this address

A progress log (sent_log.csv by default) records every address as sent/failed,
so if the run is interrupted you can simply run it again and it resumes where it
left off without double-emailing anyone.
"""
import argparse
import base64
import csv
import os
import re
import sys
import time

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

import openpyxl
import resend

# --------------------------------------------------------------------------
# Email content
# --------------------------------------------------------------------------
SUBJECT = "Thank you for joining MyOBE - and a look at our Institutional Plans"

PLAIN_BODY = """\
Dear Educator,

Thank you for registering with MyOBE - we're delighted to have you on board.

MyOBE was built to take the paperwork out of outcome-based education. From
course outcomes and question papers to lesson plans, teaching diaries, and
NBA/NAAC-ready CO-PO attainment reports, our AI helps you produce
accreditation-quality documentation in minutes instead of days.

As you explore the platform, we wanted to share something that may be useful if
you're considering MyOBE for your department or institution. We've put together
three Institutional Plans that place a shared token pool behind your entire
faculty - one account, central control, and better value at scale. Full details
are in the attached document.

    Starter       Rs 1,00,000     16,000 tokens
    Growth        Rs 2,00,000     33,000 tokens
    Enterprise    Rs 3,00,000     50,000 tokens

If you'd like to bring MyOBE to your campus, or simply have questions about how
it can fit your workflow, just reply to this email or write to us at
help@myobe.in - we'll be glad to help and can set up an institution account
within one working day.

Thank you once again for choosing MyOBE. We look forward to supporting you.

Warm regards,
The MyOBE Team
help@myobe.in | myobe.in

---
You're receiving this because you registered at myobe.in.
To stop receiving these emails, just reply with "UNSUBSCRIBE".
"""

HTML_BODY = """\
<div style="font-family:Arial,Helvetica,sans-serif;color:#1a1a2e;font-size:15px;line-height:1.6;max-width:620px;">
  <p>Dear Educator,</p>
  <p>Thank you for registering with <strong>MyOBE</strong> - we're delighted to have you on board.</p>
  <p>MyOBE was built to take the paperwork out of outcome-based education. From course
     outcomes and question papers to lesson plans, teaching diaries, and NBA/NAAC-ready
     CO-PO attainment reports, our AI helps you produce accreditation-quality
     documentation in minutes instead of days.</p>
  <p>As you explore the platform, we wanted to share something that may be useful if
     you're considering MyOBE for your <strong>department or institution</strong>.
     We've put together three <strong>Institutional Plans</strong> that place a shared
     token pool behind your entire faculty - one account, central control, and better
     value at scale. Full details are in the attached document.</p>
  <table role="presentation" cellpadding="0" cellspacing="0" style="border-collapse:collapse;margin:16px 0;font-size:14px;">
    <tr style="background:#1e2a78;color:#ffffff;">
      <th style="text-align:left;padding:8px 16px;">Plan</th>
      <th style="text-align:left;padding:8px 16px;">Investment</th>
      <th style="text-align:left;padding:8px 16px;">Token Pool</th>
    </tr>
    <tr style="background:#eef2ff;">
      <td style="padding:8px 16px;"><strong>Starter</strong></td>
      <td style="padding:8px 16px;">Rs 1,00,000</td>
      <td style="padding:8px 16px;">16,000 tokens</td>
    </tr>
    <tr>
      <td style="padding:8px 16px;"><strong>Growth</strong></td>
      <td style="padding:8px 16px;">Rs 2,00,000</td>
      <td style="padding:8px 16px;">33,000 tokens</td>
    </tr>
    <tr style="background:#e9f7ef;">
      <td style="padding:8px 16px;"><strong>Enterprise</strong></td>
      <td style="padding:8px 16px;">Rs 3,00,000</td>
      <td style="padding:8px 16px;">50,000 tokens</td>
    </tr>
  </table>
  <p>If you'd like to bring MyOBE to your campus, or simply have questions about how it
     can fit your workflow, just reply to this email or write to us at
     <a href="mailto:help@myobe.in" style="color:#4f46e5;">help@myobe.in</a> - we'll be
     glad to help and can set up an institution account within one working day.</p>
  <p>Thank you once again for choosing MyOBE. We look forward to supporting you.</p>
  <p>Warm regards,<br><strong>The MyOBE Team</strong><br>
     <a href="mailto:help@myobe.in" style="color:#4f46e5;">help@myobe.in</a> |
     <a href="https://myobe.in" style="color:#4f46e5;">myobe.in</a></p>
  <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0;">
  <p style="color:#94a3b8;font-size:12px;">You're receiving this because you registered
     at myobe.in. To stop receiving these emails, just reply with "UNSUBSCRIBE".</p>
</div>
"""

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


# --------------------------------------------------------------------------
# Excel parsing
# --------------------------------------------------------------------------
def _col_index(ws, column):
    """Resolve a --column argument (header name, letter, or 1-based number)
    to a 0-based column index. Returns None to mean 'scan every column'."""
    if column is None:
        return None
    # numeric -> 1-based column number
    if column.isdigit():
        return int(column) - 1
    # single/double letters -> spreadsheet column (A, B, ...)
    if re.fullmatch(r"[A-Za-z]{1,3}", column):
        return openpyxl.utils.column_index_from_string(column.upper()) - 1
    # otherwise treat as a header label in the first row
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for i, val in enumerate(header):
        if val and str(val).strip().lower() == column.strip().lower():
            return i
    raise SystemExit(f"Column '{column}' not found in the header row.")


def extract_emails(path, sheet=None, column=None):
    """Return a de-duplicated, order-preserving list of email addresses found
    in the workbook. By default every cell is scanned; --column narrows it."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    col = _col_index(ws, column)

    seen, emails = set(), []
    for row in ws.iter_rows(values_only=True):
        cells = [row[col]] if (col is not None and col < len(row)) else row
        for cell in cells:
            if cell is None:
                continue
            for match in EMAIL_RE.findall(str(cell)):
                key = match.strip().lower()
                if key not in seen:
                    seen.add(key)
                    emails.append(key)
    wb.close()
    return emails


# --------------------------------------------------------------------------
# Progress log (resume support)
# --------------------------------------------------------------------------
def load_done(log_path):
    """Emails already recorded as 'sent' in the log, so we never double-send."""
    done = set()
    if os.path.exists(log_path):
        with open(log_path, newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                if len(row) >= 2 and row[1] == "sent":
                    done.add(row[0].lower())
    return done


def log_result(log_path, email, status, detail=""):
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow([email, status, time.strftime("%Y-%m-%d %H:%M:%S"), detail])


# --------------------------------------------------------------------------
# Sending
# --------------------------------------------------------------------------
def build_attachment(path):
    with open(path, "rb") as f:
        content = base64.b64encode(f.read()).decode("ascii")
    return {"filename": os.path.basename(path), "content": content}


def send_one(to_email, sender, attachment):
    resend.Emails.send({
        "from": sender,
        "to": [to_email],
        "subject": SUBJECT,
        "html": HTML_BODY,
        "text": PLAIN_BODY,
        "attachments": [attachment],
    })


def main():
    ap = argparse.ArgumentParser(description="Send the MyOBE institutional-plan email individually to each recipient.")
    ap.add_argument("--excel", help="Path to the .xlsx file with recipient emails.")
    ap.add_argument("--attachment", default="MyOBE_Institutional_Plans.docx",
                    help="File to attach (default: MyOBE_Institutional_Plans.docx).")
    ap.add_argument("--column", help="Restrict email extraction to this column (header name, letter, or number).")
    ap.add_argument("--sheet", help="Worksheet name (default: the active sheet).")
    ap.add_argument("--delay", type=float, default=0.6, help="Seconds to pause between sends (default 0.6).")
    ap.add_argument("--limit", type=int, help="Send to at most this many recipients (test batch).")
    ap.add_argument("--log", default="sent_log.csv", help="Progress log; already-sent addresses are skipped on re-run.")
    ap.add_argument("--from", dest="sender", help="Override the From address (default: $SMTP_FROM).")
    ap.add_argument("--test", help="Send a single email to this address and exit (ignores the sheet).")
    ap.add_argument("--dry-run", action="store_true", help="Parse and preview only; send nothing.")
    args = ap.parse_args()

    sender = args.sender or os.environ.get("SMTP_FROM", "MyOBE <help@myobe.in>").strip()
    api_key = os.environ.get("RESEND_API_KEY", "").strip()

    # Attachment must exist regardless of mode.
    if not os.path.exists(args.attachment):
        sys.exit(f"Attachment not found: {args.attachment}")

    # Build recipient list.
    if args.test:
        recipients = [args.test.strip().lower()]
    else:
        if not args.excel:
            sys.exit("Provide --excel <file.xlsx> (or --test <email> for a single test).")
        if not os.path.exists(args.excel):
            sys.exit(f"Excel file not found: {args.excel}")
        recipients = extract_emails(args.excel, sheet=args.sheet, column=args.column)

    if not recipients:
        sys.exit("No email addresses found.")

    done = load_done(args.log)
    pending = [e for e in recipients if e not in done]
    skipped = len(recipients) - len(pending)
    if args.limit:
        pending = pending[:args.limit]

    print(f"From:        {sender}")
    print(f"Attachment:  {args.attachment}")
    print(f"Found:       {len(recipients)} unique address(es)")
    if skipped:
        print(f"Skipping:    {skipped} already in {args.log}")
    print(f"To send now: {len(pending)}")

    if args.dry_run:
        print("\n-- DRY RUN, nothing sent. Recipients: --")
        for e in pending:
            print("  ", e)
        return

    if not api_key:
        sys.exit("RESEND_API_KEY not set. Export it or add it to your .env, then retry.")
    resend.api_key = api_key
    attachment = build_attachment(args.attachment)

    sent = failed = 0
    for i, email in enumerate(pending, 1):
        try:
            send_one(email, sender, attachment)
            sent += 1
            log_result(args.log, email, "sent")
            print(f"[{i}/{len(pending)}] sent    -> {email}")
        except Exception as e:
            failed += 1
            log_result(args.log, email, "failed", str(e))
            print(f"[{i}/{len(pending)}] FAILED  -> {email}: {e}")
        if i < len(pending):
            time.sleep(args.delay)

    print(f"\nDone. Sent: {sent}  Failed: {failed}  (log: {args.log})")
    if failed:
        print("Re-run the same command to retry only the failed/remaining addresses.")


if __name__ == "__main__":
    main()

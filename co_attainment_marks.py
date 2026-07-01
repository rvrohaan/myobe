"""Marks-based CO attainment (Method-2 Tier-I) for Module 4.

This module is self-contained (no Flask imports) and provides three things:

  1. build_template_xlsx()        - a styled .xlsx the user can download and fill.
  2. parse_marks_workbook(path)   - read a filled .xlsx / legacy .xls into a dict.
  3. compute_marks_attainment()   - turn parsed marks into CO attainment levels.

Methodology (faithful to the institution's "Method-2 Tier-I" reference sheet):

  * Every exam question is tagged to a CO with a max-marks value, for two
    components: CIE (internal tests) and SEE (semester-end exam).
  * Per student, per CO:  CO%% = sum(marks on that CO's questions) / sum(max).
  * Per CO across the class:
        assessment%% = (students scoring CO%% >= target) / students-present.
  * assessment%% maps to a 0-3 level via component thresholds
        CIE: >=50/60/70 -> L1/L2/L3       SEE: >=60/70/80 -> L1/L2/L3
  * Combined CO attainment = 0.2 * CIE_level + 0.8 * SEE_level   (0-3 scale).

All user-visible strings are plain ASCII on purpose.
"""

import io


# -- Defaults (match the reference sheet) -------------------------------------

DEFAULTS = {
    "target_pct":     60,
    "ia_thresholds":  {"l1": 60, "l2": 70, "l3": 80},   # CIE
    "see_thresholds": {"l1": 60, "l2": 70, "l3": 80},   # SEE
    "cie_weight":     20,
    "see_weight":     80,
}

# Column-A row labels the parser anchors on (matched case-insensitively).
_LABEL_COMPONENT = "component"
_LABEL_QUESTION  = "question"
_LABEL_CO        = "co"
_LABEL_MAXMARKS  = "max marks"
_LABEL_BLOOM     = "bloom"

# Optional per-CO maximum (denominator) row. Lets reference-style sheets with
# internal choice / multiple tests state each CO's attainable max directly,
# instead of summing question max-marks (which assumes every question is attempted).
_COMAX_LABELS = {"co max", "co maximum", "co max marks", "co total",
                 "co total marks", "co maximum marks"}

_HEADER_LABELS = {
    _LABEL_COMPONENT, _LABEL_QUESTION, _LABEL_CO, _LABEL_MAXMARKS, _LABEL_BLOOM,
    "blooms level", "maxmarks", "max-marks", "student id", "usn",
} | _COMAX_LABELS


# -- Small helpers -------------------------------------------------------------

def _norm_label(v):
    return str(v or "").strip().lower()


def _norm_co(v):
    """Normalize a CO label so 'CO 1', 'co1', 'CO1' all become 'CO1'."""
    s = str(v or "").strip().upper().replace(" ", "")
    return s


def _to_float(v):
    """Best-effort numeric parse; '' / None -> None, junk -> None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _component(v):
    s = str(v or "").strip().upper()
    if s.startswith("SEE"):
        return "SEE"
    return "CIE"   # default / blank -> CIE


def _component_marker(v):
    """Classify a free-text section title as CIE / SEE, or None if it isn't one.

    Recognizes the section headers a reference-style sheet uses to label its
    blocks (e.g. "Test Marks" over the internal tests, "Semester End
    Examination" over the final exam) in addition to plain "CIE" / "SEE".
    """
    s = str(v or "").strip().lower()
    if not s:
        return None
    if (s == "see" or "semester end" in s or "semester-end" in s
            or "end examination" in s or "end exam" in s or "sem end" in s):
        return "SEE"
    if (s == "cie" or "test marks" in s or "internal" in s or s == "ia"
            or "continuous internal" in s or "cia" in s):
        return "CIE"
    return None


def _summary_kind(v):
    """Classify a computed-summary section title, or None if it isn't one.

    Reference sheets interleave computed blocks between the CIE and SEE question
    blocks. "PCT" marks per-student CO percentages we can use directly (e.g.
    "CO IN PERCENTAGE", "CO Percentage"); "SKIP" marks other summaries that must
    not be read as questions (e.g. "CO Total Marks", "Q No. to be entered").
    """
    s = str(v or "").strip().lower()
    if not s:
        return None
    if "percentage" in s or s.endswith("percent") or " percent" in s:
        return "PCT"
    # Only true summary-block titles SKIP a run of columns. Stray annotations
    # like "Q No. to be entered" sit on real question columns and must NOT
    # forward-fill SKIP across them.
    if "co total" in s or "total marks" in s:
        return "SKIP"
    return None


def _build_component_map(marks_grid, header_rows_idx, n_cols, comp_row):
    """Resolve each column's kind and component region.

    Returns (col_kind, region, source):
      col_kind[c] in {'CIE','SEE','SKIP','PCT'} - how the column is used:
        CIE/SEE  -> raw question column for that component
        SKIP     -> computed summary (CO totals, "Q No.") - ignored
        PCT      -> precomputed per-student CO percentage - used directly
      region[c]  in {'CIE','SEE'} - which exam block the column physically sits
        in (forward-fill of CIE/SEE titles only); used to assign a component to
        PCT and question columns.
      source     in {'component_row','section_markers','assumed_cie'}.
    """
    kind_markers = []    # (col, 'CIE'|'SEE'|'SKIP'|'PCT')
    region_markers = []  # (col, 'CIE'|'SEE')
    source = "assumed_cie"

    if comp_row is not None:
        for c in range(1, n_cols):
            val = comp_row[c] if c < len(comp_row) else None
            if val is not None and str(val).strip():
                comp = _component(val)
                kind_markers.append((c, comp))
                region_markers.append((c, comp))
        if kind_markers:
            source = "component_row"
    else:
        scan = set(header_rows_idx)
        first = min(header_rows_idx) if header_rows_idx else 0
        scan.update(range(0, first))   # section titles often sit above the headers
        for r in sorted(scan):
            if r >= len(marks_grid):
                continue
            row = marks_grid[r]
            for c in range(0, min(len(row), n_cols)):
                summ = _summary_kind(row[c])
                if summ:
                    kind_markers.append((c, summ))
                    continue
                comp = _component_marker(row[c])
                if comp:
                    kind_markers.append((c, comp))
                    region_markers.append((c, comp))
        if region_markers:
            source = "section_markers"

    def _fill(markers, default):
        markers = sorted(markers)
        out, cur, mi = {}, default, 0
        for c in range(0, n_cols):
            while mi < len(markers) and markers[mi][0] <= c:
                cur = markers[mi][1]
                mi += 1
            out[c] = cur
        return out

    col_kind = _fill(kind_markers, "CIE")
    region   = _fill(region_markers, "CIE")
    return col_kind, region, source


# -- Template builder ----------------------------------------------------------

def build_template_xlsx():
    """Return a BytesIO holding a ready-to-fill .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    green   = "059669"
    dgreen  = "D1FAE5"
    grey    = "6B7280"
    hdr_fill   = PatternFill("solid", fgColor=dgreen)
    label_fill = PatternFill("solid", fgColor="F3F4F6")
    bold       = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    thin       = Side(style="thin", color="D1D5DB")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center")

    # ---- Instructions sheet --------------------------------------------------
    ins = wb.active
    ins.title = "Instructions"
    ins.column_dimensions["A"].width = 100
    lines = [
        ("How to fill this Student Marks workbook", True),
        ("", False),
        ("This workbook computes CO attainment from real student marks using the", False),
        ("Method-2 Tier-I methodology (CIE internal tests + SEE semester exam).", False),
        ("", False),
        ("1. Go to the 'Marks' sheet.", True),
        ("2. The first rows are the question definition (do NOT delete them):", False),
        ("     - Component : CIE for internal tests, SEE for the semester-end exam.", False),
        ("     - Question  : any label for the question (e.g. 1-A, 2-B).", False),
        ("     - CO        : the Course Outcome the question assesses (CO1, CO2, ...).", False),
        ("     - Max Marks : the maximum marks for that question.", False),
        ("     - Bloom     : optional Bloom level (L1..L6). Not used in the maths.", False),
        ("     - CO Max    : OPTIONAL row (see below). Leave blank for normal use.", False),
        ("3. Add or remove question columns as needed. Keep Component/CO/Max Marks filled.", False),
        ("4. Below the header rows, enter one student per row:", False),
        ("     - Column A  : the student USN / ID.", False),
        ("     - Other cols: the marks the student scored on that question.", False),
        ("     - Leave a cell blank if the student did not attempt that question.", False),
        ("     - Leave ALL of a component's cells blank for a student who was absent.", False),
        ("5. The 'Config' sheet holds the target % and attainment thresholds.", False),
        ("   Change them only if your institution uses different rules.", False),
        ("6. Save the file and upload it back into Module 4.", False),
        ("", False),
        ("When to use the optional 'CO Max' row:", True),
        ("Normal case - every student attempts every listed question: leave the", False),
        ("'CO Max' row BLANK. The CO % is marks / (sum of that CO's question marks).", False),
        ("Internal choice - a student answers only some of a CO's questions across", False),
        ("several tests: fill 'CO Max' with each CO's attainable maximum, entered", False),
        ("under any one of that CO's columns. The CO % is then marks / CO Max.", False),
        ("CIE and SEE blocks can carry different CO Max values.", False),
        ("", False),
        ("Attainment is the percentage of students scoring at or above the target", False),
        ("in each CO, converted to a level, then combined as 20% CIE + 80% SEE.", False),
    ]
    for i, (text, is_bold) in enumerate(lines, start=1):
        c = ins.cell(row=i, column=1, value=text)
        c.font = Font(bold=True, size=13, color=green) if (i == 1) else (bold if is_bold else Font(color="111827"))

    # ---- Marks sheet ---------------------------------------------------------
    ms = wb.create_sheet("Marks")

    # Example question definition: 2 CIE + 2 SEE questions for each of CO1..CO5.
    cos = ["CO1", "CO2", "CO3", "CO4", "CO5"]
    questions = []   # (component, qlabel, co, maxmarks, bloom)
    for idx, co in enumerate(cos, start=1):
        questions.append(("CIE", f"{idx}-A", co, 10, "L2"))
        questions.append(("CIE", f"{idx}-B", co, 5,  "L3"))
    for idx, co in enumerate(cos, start=1):
        questions.append(("SEE", f"{idx}-A", co, 10, "L2"))
        questions.append(("SEE", f"{idx}-B", co, 5,  "L3"))

    # Header rows (label in column A, values from column B onward). The
    # "CO Max" row is optional: leave it blank unless students attempt only a
    # subset of each CO's questions (internal choice across multiple tests).
    header_rows = [
        ("Component",        [q[0] for q in questions],      False),
        ("Question",         [q[1] for q in questions],      False),
        ("CO",               [q[2] for q in questions],      False),
        ("Max Marks",        [q[3] for q in questions],      False),
        ("CO Max (optional)", [""] * len(questions),         True),
        ("Bloom",            [q[4] for q in questions],      False),
    ]
    for r, (label, vals, optional) in enumerate(header_rows, start=1):
        lc = ms.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, italic=optional, color=(grey if optional else "111827"))
        lc.fill = label_fill
        lc.border = border
        if optional:
            continue   # leave the value cells blank for the user to fill if needed
        for cidx, v in enumerate(vals, start=2):
            cell = ms.cell(row=r, column=cidx, value=v)
            cell.fill = hdr_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border

    # Example student rows (deterministic, illustrative marks).
    data_start = len(header_rows) + 1
    sample = [
        ("CS001", [10, 5, 8, 4, 9, 5, 7, 3, 10, 4,  9, 4, 8, 5, 9, 4, 6, 3, 8, 5]),
        ("CS002", [8,  4, 9, 5, 7, 2, 10, 5, 6, 3,  10, 5, 7, 4, 8, 5, 9, 3, 7, 4]),
        ("CS003", [9,  3, 6, 2, 8, 4, 5, 1, 9, 5,   8, 4, 9, 5, 6, 2, 10, 5, 8, 3]),
        ("CS004", [7,  5, 10, 4, 6, 3, 8, 4, 7, 2,  9, 3, 8, 4, 7, 3, 8, 4, 9, 5]),
        ("CS005", [10, 4, 7, 3, 9, 5, 6, 2, 8, 4,   7, 5, 10, 4, 9, 5, 7, 2, 8, 4]),
    ]
    for r, (usn, marks) in enumerate(sample, start=data_start):
        ms.cell(row=r, column=1, value=usn).border = border
        for cidx, m in enumerate(marks, start=2):
            cell = ms.cell(row=r, column=cidx, value=m)
            cell.alignment = center
            cell.border = border

    ms.column_dimensions["A"].width = 18
    ms.freeze_panes = "B7"

    # ---- Config sheet --------------------------------------------------------
    cfg = wb.create_sheet("Config")
    cfg.column_dimensions["A"].width = 26
    cfg.column_dimensions["B"].width = 12
    cfg_rows = [
        ("Setting", "Value"),
        ("Target %", DEFAULTS["target_pct"]),
        ("CIE Level 1 >=", DEFAULTS["ia_thresholds"]["l1"]),
        ("CIE Level 2 >=", DEFAULTS["ia_thresholds"]["l2"]),
        ("CIE Level 3 >=", DEFAULTS["ia_thresholds"]["l3"]),
        ("SEE Level 1 >=", DEFAULTS["see_thresholds"]["l1"]),
        ("SEE Level 2 >=", DEFAULTS["see_thresholds"]["l2"]),
        ("SEE Level 3 >=", DEFAULTS["see_thresholds"]["l3"]),
        ("CIE Weight %", DEFAULTS["cie_weight"]),
        ("SEE Weight %", DEFAULTS["see_weight"]),
    ]
    for r, (k, v) in enumerate(cfg_rows, start=1):
        kc = cfg.cell(row=r, column=1, value=k)
        vc = cfg.cell(row=r, column=2, value=v)
        if r == 1:
            kc.font = white_bold; kc.fill = PatternFill("solid", fgColor=green)
            vc.font = white_bold; vc.fill = PatternFill("solid", fgColor=green)
        else:
            kc.font = bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# -- Workbook parsing ----------------------------------------------------------

def _read_grid(path, ext):
    """Return the first usable sheet as a list-of-rows grid of cell values.

    Picks the sheet named 'Marks' when present, else the first sheet that has a
    'CO' / 'Max Marks' header. Also returns the Config sheet rows (or None).
    """
    ext = (ext or "").lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        sheets = {s.name: [[s.cell_value(r, c) for c in range(s.ncols)]
                           for r in range(s.nrows)] for s in wb.sheets()}
    else:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        sheets = {ws.title: [list(row) for row in ws.iter_rows(values_only=True)]
                  for ws in wb.worksheets}

    def _has_headers(grid):
        labels = {_norm_label(row[0]) for row in grid if row}
        return _LABEL_CO in labels and _LABEL_MAXMARKS in labels

    marks_grid = None
    for name, grid in sheets.items():
        if name.strip().lower() == "marks" and _has_headers(grid):
            marks_grid = grid
            break
    if marks_grid is None:
        for grid in sheets.values():
            if _has_headers(grid):
                marks_grid = grid
                break
    if marks_grid is None:
        raise ValueError(
            "Could not find a 'CO' and 'Max Marks' header in the workbook. "
            "Please download the sample template, fill in your marks, and upload it.")

    config_grid = next((g for n, g in sheets.items()
                        if n.strip().lower() == "config"), None)
    return marks_grid, config_grid


def _parse_config(config_grid):
    cfg = {
        "target_pct":     DEFAULTS["target_pct"],
        "ia_thresholds":  dict(DEFAULTS["ia_thresholds"]),
        "see_thresholds": dict(DEFAULTS["see_thresholds"]),
        "cie_weight":     DEFAULTS["cie_weight"],
        "see_weight":     DEFAULTS["see_weight"],
    }
    if not config_grid:
        return cfg
    kv = {}
    for row in config_grid:
        if not row or len(row) < 2:
            continue
        kv[_norm_label(row[0])] = _to_float(row[1])

    def _set(key, dst, *path):
        val = kv.get(key)
        if val is not None:
            ref = cfg
            for p in path[:-1]:
                ref = ref[p]
            ref[path[-1]] = val

    _set("target %", cfg, "target_pct")
    _set("cie level 1 >=", cfg, "ia_thresholds", "l1")
    _set("cie level 2 >=", cfg, "ia_thresholds", "l2")
    _set("cie level 3 >=", cfg, "ia_thresholds", "l3")
    _set("see level 1 >=", cfg, "see_thresholds", "l1")
    _set("see level 2 >=", cfg, "see_thresholds", "l2")
    _set("see level 3 >=", cfg, "see_thresholds", "l3")
    _set("cie weight %", cfg, "cie_weight")
    _set("see weight %", cfg, "see_weight")
    return cfg


def parse_marks_workbook(path, ext):
    """Parse a filled marks workbook into:

        {questions:[{component, co, max_marks, col}], students:[{id, marks}],
         config:{...}}

    Raises ValueError with a clear message on malformed input.
    """
    marks_grid, config_grid = _read_grid(path, ext)

    # Locate header rows by their column-A label.
    rows_by_label = {}
    for r, row in enumerate(marks_grid):
        if not row:
            continue
        lbl = _norm_label(row[0])
        if lbl in (_LABEL_COMPONENT, _LABEL_QUESTION, _LABEL_CO,
                   _LABEL_MAXMARKS, _LABEL_BLOOM, "blooms level"):
            key = "bloom" if lbl in (_LABEL_BLOOM, "blooms level") else lbl.replace(" ", "")
            rows_by_label.setdefault(key, r)
        elif lbl in _COMAX_LABELS or lbl.startswith("co max") or lbl.startswith("co maximum"):
            rows_by_label.setdefault("comax", r)

    if "co" not in rows_by_label or "maxmarks" not in rows_by_label:
        raise ValueError("Missing the 'CO' or 'Max Marks' header row. "
                         "Please download the sample template, fill in your marks, and upload it.")

    co_row    = marks_grid[rows_by_label["co"]]
    max_row   = marks_grid[rows_by_label["maxmarks"]]
    comp_row  = marks_grid[rows_by_label["component"]] if "component" in rows_by_label else None
    comax_row = marks_grid[rows_by_label["comax"]] if "comax" in rows_by_label else None

    n_cols = max(len(co_row), len(max_row),
                 *(len(marks_grid[r]) for r in rows_by_label.values()))

    # Resolve each column's kind (CIE/SEE/SKIP/PCT) and physical region (CIE/SEE)
    # from an explicit Component row or section-title markers (reference layout).
    col_kind, region, component_source = _build_component_map(
        marks_grid, list(rows_by_label.values()), n_cols, comp_row)

    # Raw question column: a CO label + positive max-marks, not summary/percentage.
    questions = []
    for c in range(1, n_cols):
        if col_kind.get(c) in ("SKIP", "PCT"):
            continue
        co  = _norm_co(co_row[c]) if c < len(co_row) else ""
        mx  = _to_float(max_row[c]) if c < len(max_row) else None
        if not co or co in ("", "CO") or mx is None or mx <= 0:
            continue
        questions.append({"component": region.get(c, "CIE"),
                          "co": co, "max_marks": mx, "col": c})

    # Precomputed per-student CO percentage columns ("CO IN PERCENTAGE" /
    # "CO Percentage"). When present they are the most reliable source, since the
    # sheet has already accounted for internal choice across multiple tests.
    pct_columns = []   # [{component, co, col}]
    for c in range(1, n_cols):
        if col_kind.get(c) != "PCT":
            continue
        co = _norm_co(co_row[c]) if c < len(co_row) else ""
        if co and co not in ("", "CO"):
            pct_columns.append({"component": region.get(c, "CIE"), "co": co, "col": c})

    # Optional per-(component, CO) max override from a "CO Max" row.
    co_max_override = {}
    if comax_row is not None:
        for c in range(1, n_cols):
            if col_kind.get(c) in ("SKIP", "PCT"):
                continue
            co = _norm_co(co_row[c]) if c < len(co_row) else ""
            if not co or co in ("", "CO"):
                continue
            val = _to_float(comax_row[c]) if c < len(comax_row) else None
            if val is not None and val > 0:
                co_max_override.setdefault((region.get(c, "CIE"), co), val)

    if not questions and not pct_columns:
        raise ValueError("No question or CO-percentage columns found. "
                         "Please download the sample template, fill in your marks, and upload it.")

    # Student rows: everything after the last header row, with a non-empty,
    # non-label column-A value. Read both raw question columns and percentage
    # columns into the per-student marks dict (keyed by column index).
    read_cols = [q["col"] for q in questions] + [p["col"] for p in pct_columns]
    last_header = max(rows_by_label.values())
    students = []
    for r in range(last_header + 1, len(marks_grid)):
        row = marks_grid[r]
        if not row:
            continue
        sid = str(row[0]).strip() if row[0] is not None else ""
        if sid == "" or _norm_label(sid) in _HEADER_LABELS:
            continue
        marks = {c: (_to_float(row[c]) if c < len(row) else None) for c in read_cols}
        students.append({"id": sid, "marks": marks})

    if not students:
        raise ValueError("No student rows found below the header. "
                         "Add one student per row with their marks.")

    return {
        "questions":        questions,
        "pct_columns":      pct_columns,
        "students":         students,
        "config":           _parse_config(config_grid),
        "co_max_override":  co_max_override,
        "component_source": component_source,
    }


# -- Attainment computation ----------------------------------------------------

def _level(assessment_pct, thr):
    if assessment_pct >= thr["l3"]:
        return 3
    if assessment_pct >= thr["l2"]:
        return 2
    if assessment_pct >= thr["l1"]:
        return 1
    return 0


def _assess(per_student_pct, target, thr, total):
    """Given each present student's CO% list, return assessment summary."""
    if total == 0:
        return {"assessment_pct": 0.0, "level": 0, "students_present": 0}
    above = sum(1 for p in per_student_pct if p >= target)
    ap = above / total * 100.0
    return {"assessment_pct": round(ap, 2), "level": _level(ap, thr),
            "students_present": total}


def _component_attainment(questions, students, cos, target, thr, component,
                          co_max_override=None, pct_columns=None):
    """Return {co -> {assessment_pct, level, students_present}} for one component.

    Uses precomputed per-student CO% columns when available (most reliable for
    reference sheets); otherwise computes CO% from raw marks, with an optional
    per-CO max override, falling back to the sum of question max-marks.
    """
    co_max_override = co_max_override or {}
    pct_cols = [p for p in (pct_columns or []) if p["component"] == component]
    out = {}

    # -- Precomputed percentage path -----------------------------------------
    if pct_cols:
        col_by_co = {}
        for p in pct_cols:
            col_by_co.setdefault(p["co"], p["col"])
        cols = list(col_by_co.values())
        present = [s for s in students
                   if any(s["marks"].get(c) is not None for c in cols)]
        total = len(present)
        for co in cos:
            col = col_by_co.get(co)
            if col is None:
                out[co] = {"assessment_pct": 0.0, "level": 0, "students_present": total}
                continue
            vals = [s["marks"].get(col) for s in present if s["marks"].get(col) is not None]
            out[co] = _assess(vals, target, thr, total)
        return out

    # -- Raw marks path ------------------------------------------------------
    qs = [q for q in questions if q["component"] == component]
    if not qs:
        return out
    cols = [q["col"] for q in qs]
    present = [s for s in students
               if any(s["marks"].get(c) is not None for c in cols)]
    total = len(present)
    for co in cos:
        co_qs  = [q for q in qs if q["co"] == co]
        # Prefer an explicit per-CO max (handles internal choice / multiple
        # tests); otherwise sum the CO's question max-marks.
        co_max = co_max_override.get((component, co)) or sum(q["max_marks"] for q in co_qs)
        if co_max <= 0 or total == 0:
            out[co] = {"assessment_pct": 0.0, "level": 0, "students_present": total}
            continue
        vals = [sum((s["marks"].get(q["col"]) or 0.0) for q in co_qs) / co_max * 100.0
                for s in present]
        out[co] = _assess(vals, target, thr, total)
    return out


def compute_marks_attainment(parsed):
    """Compute Method-2 Tier-I CO attainment from a parsed workbook.

    Returns a payload shaped to flow into the Module-4 report alongside the
    AI-estimated `coatt` blob.
    """
    questions   = parsed["questions"]
    pct_columns = parsed.get("pct_columns") or []
    students    = parsed["students"]
    cfg         = parsed["config"]

    target  = cfg["target_pct"]
    ia_thr  = cfg["ia_thresholds"]
    see_thr = cfg["see_thresholds"]
    cie_w   = cfg["cie_weight"]
    see_w   = cfg["see_weight"]

    # Preserve CO order of first appearance (raw questions first, then any CO
    # that only appears in percentage columns).
    cos = []
    for item in list(questions) + list(pct_columns):
        if item["co"] not in cos:
            cos.append(item["co"])

    has_cie = (any(q["component"] == "CIE" for q in questions)
               or any(p["component"] == "CIE" for p in pct_columns))
    has_see = (any(q["component"] == "SEE" for q in questions)
               or any(p["component"] == "SEE" for p in pct_columns))

    co_max_override = parsed.get("co_max_override") or {}
    cie = _component_attainment(questions, students, cos, target, ia_thr, "CIE",
                                co_max_override, pct_columns)
    see = _component_attainment(questions, students, cos, target, see_thr, "SEE",
                                co_max_override, pct_columns)

    co_results = []
    for co in cos:
        ia = cie.get(co, {"assessment_pct": 0.0, "level": 0})
        se = see.get(co, {"assessment_pct": 0.0, "level": 0})

        if has_cie and has_see:
            combined = (cie_w / 100.0) * ia["level"] + (see_w / 100.0) * se["level"]
        elif has_see:
            combined = float(se["level"])
        else:
            combined = float(ia["level"])

        co_results.append({
            "co":        co,
            "ia_pct":    ia["assessment_pct"] if has_cie else None,
            "see_pct":   se["assessment_pct"] if has_see else None,
            "ia_level":  ia["level"] if has_cie else None,
            "see_level": se["level"] if has_see else None,
            "level":     round(combined, 2),
            "pct":       round(combined / 3.0 * 100.0, 2),
        })

    n_cie = max((cie.get(c, {}).get("students_present", 0) for c in cos), default=0)
    n_see = max((see.get(c, {}).get("students_present", 0) for c in cos), default=0)

    return {
        "source":    "marks",
        "coResults": co_results,
        "config":    cfg,
        "summary": {
            "n_students_cie":   n_cie,
            "n_students_see":   n_see,
            "n_cos":            len(cos),
            "n_questions":      len(questions),
            "target_pct":       target,
            "has_cie":          has_cie,
            "has_see":          has_see,
            "cie_weight":       cie_w,
            "see_weight":       see_w,
            "component_source": parsed.get("component_source", "assumed_cie"),
        },
    }
